"""
Multiple Regression Engine – Minitab-Style Reports
Supports: OLS, Ridge, Lasso, ElasticNet, PLS
"""

import numpy as np
import pandas as pd
from scipy import stats
from scipy.stats import norm, shapiro
from itertools import combinations
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from typing import Dict, List, Optional, Tuple


# =============================================================================
# 1. FEATURE ENGINEERING
# =============================================================================

def generate_extended_features(
    X: np.ndarray,
    names: List[str],
    include_squares: bool = True,
    include_interactions: bool = True,
) -> Tuple[np.ndarray, List[str], List[bool]]:
    """
    Generate polynomial (squared) and interaction features.
    Returns extended X matrix, feature names, and boolean mask
    indicating which features are higher-order (for visual styling).
    """
    cols = [X[:, i] for i in range(X.shape[1])]
    out_names = list(names)
    is_higher = [False] * len(names)

    if include_squares:
        for i, n in enumerate(names):
            cols.append(X[:, i] ** 2)
            out_names.append(f"{n}^2")
            is_higher.append(True)

    if include_interactions:
        for i, j in combinations(range(len(names)), 2):
            cols.append(X[:, i] * X[:, j])
            out_names.append(f"{names[i]}*{names[j]}")
            is_higher.append(True)

    X_ext = np.column_stack(cols)
    return X_ext, out_names, is_higher


# =============================================================================
# 2. OLS STATISTICS (pure numpy/scipy – no statsmodels dependency)
# =============================================================================

def compute_ols_statistics(
    X: np.ndarray,
    y: np.ndarray,
    feature_names: List[str],
) -> Dict:
    """Full OLS statistics with p-values, R², F-test."""
    n, p = X.shape
    X_design = np.column_stack([np.ones(n), X])
    k = p + 1  # parameters including intercept

    if n <= k:
        raise ValueError(
            f"Not enough observations ({n}) for {p} predictors. "
            f"Need at least {p + 2} observations."
        )

    # OLS via least squares
    beta, res, rank, sv = np.linalg.lstsq(X_design, y, rcond=None)
    y_pred = X_design @ beta
    residuals = y - y_pred

    ss_res = np.sum(residuals ** 2)
    ss_tot = np.sum((y - y.mean()) ** 2)
    ss_reg = ss_tot - ss_res

    r_sq = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0
    r_sq_adj = 1.0 - (1.0 - r_sq) * (n - 1) / (n - k) if n > k else 0.0
    mse = ss_res / (n - k)
    s = np.sqrt(mse)

    # Coefficient standard errors and p-values
    try:
        cov = mse * np.linalg.inv(X_design.T @ X_design)
        se = np.sqrt(np.diag(cov))
    except np.linalg.LinAlgError:
        cov = mse * np.linalg.pinv(X_design.T @ X_design)
        se = np.sqrt(np.abs(np.diag(cov)))

    t_vals = beta / se
    p_vals = 2.0 * (1.0 - stats.t.cdf(np.abs(t_vals), n - k))

    # Overall F-test
    if p > 0 and mse > 0:
        f_stat = (ss_reg / p) / mse
        f_p = 1.0 - stats.f.cdf(f_stat, p, n - k)
    else:
        f_stat, f_p = 0.0, 1.0

    return {
        "coefficients": beta,
        "se": se,
        "t_values": t_vals,
        "p_values": p_vals,
        "feature_names": ["Constant"] + list(feature_names),
        "r_squared": r_sq,
        "r_squared_adj": r_sq_adj,
        "f_statistic": f_stat,
        "f_p_value": f_p,
        "residuals": residuals,
        "y_pred": y_pred,
        "mse": mse,
        "s": s,
        "n": n,
        "p": p,
    }


# =============================================================================
# 3. FORWARD STEPWISE MODEL BUILDING
# =============================================================================

def forward_stepwise(
    X: np.ndarray,
    y: np.ndarray,
    feature_names: List[str],
    is_higher: Optional[List[bool]] = None,
    alpha_enter: float = 0.25,
) -> Tuple[List[Dict], List[int]]:
    """
    Forward stepwise selection.
    Returns (steps_list, selected_indices).
    """
    n = len(y)
    n_features = X.shape[1]
    if is_higher is None:
        is_higher = [False] * n_features

    selected: List[int] = []
    remaining = list(range(n_features))
    steps: List[Dict] = []

    for step_num in range(n_features):
        best_p = 1.0
        best_idx = None

        for idx in remaining:
            candidate = selected + [idx]
            X_cand = X[:, candidate]
            if n <= len(candidate) + 1:
                continue
            try:
                st = compute_ols_statistics(X_cand, y, [feature_names[i] for i in candidate])
                p_val = st["p_values"][-1]  # p-value of newly added term
                if p_val < best_p:
                    best_p = p_val
                    best_idx = idx
            except Exception:
                continue

        if best_idx is None:
            break

        # Always add if below threshold, or if it's a linear term
        if best_p > alpha_enter and is_higher[best_idx]:
            break

        selected.append(best_idx)
        remaining.remove(best_idx)

        X_sel = X[:, selected]
        cur = compute_ols_statistics(X_sel, y, [feature_names[i] for i in selected])

        steps.append({
            "step": step_num + 1,
            "change": f"Add {feature_names[best_idx]}",
            "term": feature_names[best_idx],
            "step_p": best_p,
            "final_p": None,  # filled below
            "r_squared_adj": cur["r_squared_adj"] * 100.0,
            "is_higher": is_higher[best_idx],
        })

    # Compute final p-values
    if selected:
        final = compute_ols_statistics(X[:, selected], y,
                                       [feature_names[i] for i in selected])
        for i, step in enumerate(steps):
            step["final_p"] = final["p_values"][i + 1]  # skip intercept

    return steps, selected


# =============================================================================
# 4. MODEL FITTING (multiple model types)
# =============================================================================

def fit_model(
    X_train: np.ndarray,
    y_train: np.ndarray,
    feature_names: List[str],
    model_type: str = "OLS",
    alpha: float = 1.0,
    l1_ratio: float = 0.5,
    n_components: int = 2,
) -> Dict:
    """
    Fit a regression model and return unified result dict.
    model_type: 'OLS', 'Ridge', 'Lasso', 'ElasticNet', 'PLS'
    """
    from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet
    from sklearn.cross_decomposition import PLSRegression
    from sklearn.preprocessing import StandardScaler

    n, p = X_train.shape

    if model_type == "OLS":
        model = LinearRegression()
    elif model_type == "Ridge":
        model = Ridge(alpha=alpha)
    elif model_type == "Lasso":
        model = Lasso(alpha=alpha, max_iter=10000)
    elif model_type == "ElasticNet":
        model = ElasticNet(alpha=alpha, l1_ratio=l1_ratio, max_iter=10000)
    elif model_type == "PLS":
        nc = min(n_components, p, n - 1)
        model = PLSRegression(n_components=nc)
    else:
        raise ValueError(f"Unknown model type: {model_type}")

    model.fit(X_train, y_train)

    if model_type == "PLS":
        y_pred = model.predict(X_train).ravel()
        coefs = model.coef_.ravel()
        intercept = model.intercept_.item() if hasattr(model.intercept_, 'item') else float(model.intercept_)
    else:
        y_pred = model.predict(X_train)
        coefs = model.coef_
        intercept = model.intercept_

    residuals = y_train - y_pred
    ss_res = np.sum(residuals ** 2)
    ss_tot = np.sum((y_train - y_train.mean()) ** 2)
    r_sq = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0
    k = p + 1
    r_sq_adj = 1.0 - (1.0 - r_sq) * (n - 1) / (n - k) if n > k else 0.0
    mse = ss_res / max(n - k, 1)

    # OLS p-values (always compute for reference)
    ols = compute_ols_statistics(X_train, y_train, feature_names)

    all_coefs = np.concatenate([[intercept], coefs])

    return {
        "model": model,
        "model_type": model_type,
        "coefficients": all_coefs,
        "feature_names": ["Constant"] + list(feature_names),
        "r_squared": r_sq,
        "r_squared_adj": r_sq_adj,
        "s": np.sqrt(mse),
        "mse": mse,
        "y_pred": y_pred,
        "residuals": residuals,
        "n": n,
        "p": p,
        "ols_stats": ols,  # full OLS statistics for p-values etc.
    }


# =============================================================================
# 5. DIAGNOSTIC FUNCTIONS
# =============================================================================

def compute_incremental_impact(
    X: np.ndarray,
    y: np.ndarray,
    feature_names: List[str],
    original_names: List[str],
    selected_indices: List[int],
) -> Dict[str, float]:
    """
    For each original X variable, compute the increase in R² it provides.
    Computed by removing ALL terms involving that variable and comparing R².
    """
    if not selected_indices:
        return {n: 0.0 for n in original_names}

    X_sel = X[:, selected_indices]
    sel_names = [feature_names[i] for i in selected_indices]
    full_stats = compute_ols_statistics(X_sel, y, sel_names)
    full_r2 = full_stats["r_squared"]

    impacts = {}
    for orig in original_names:
        # Find indices of terms NOT involving this variable
        keep_mask = []
        for j, idx in enumerate(selected_indices):
            name = feature_names[idx]
            involves = orig in name.replace("^2", "").replace("*", " ").split()
            # More robust: check if the original name appears in the feature name
            if orig == name or orig in name.split("*") or name == f"{orig}^2":
                involves = True
            else:
                involves = False
            keep_mask.append(not involves)

        kept = [selected_indices[j] for j in range(len(selected_indices)) if keep_mask[j]]

        if not kept:
            impacts[orig] = full_r2 * 100.0
        else:
            X_reduced = X[:, kept]
            reduced_names = [feature_names[i] for i in kept]
            try:
                red_stats = compute_ols_statistics(X_reduced, y, reduced_names)
                impacts[orig] = (full_r2 - red_stats["r_squared"]) * 100.0
            except Exception:
                impacts[orig] = 0.0

    return impacts


def compute_x_regressed_on_others(
    X_orig: np.ndarray,
    original_names: List[str],
) -> Dict[str, float]:
    """
    For each Xi, regress Xi on all other Xj.
    Returns R² for each — high R² indicates multicollinearity.
    """
    result = {}
    p = X_orig.shape[1]
    if p < 2:
        return {n: 0.0 for n in original_names}

    for i in range(p):
        others = [j for j in range(p) if j != i]
        X_others = X_orig[:, others]
        xi = X_orig[:, i]
        try:
            ols = compute_ols_statistics(X_others, xi,
                                         [original_names[j] for j in others])
            result[original_names[i]] = ols["r_squared"] * 100.0
        except Exception:
            result[original_names[i]] = 0.0
    return result


def detect_unusual_data(
    residuals: np.ndarray,
    X: np.ndarray,
    threshold_resid: float = 2.0,
) -> Dict:
    """Detect large residuals and high-leverage points."""
    n, p = X.shape
    X_design = np.column_stack([np.ones(n), X])

    # Standardized residuals
    s = np.std(residuals, ddof=p + 1)
    std_resid = residuals / s if s > 0 else np.zeros_like(residuals)
    large_resid = np.where(np.abs(std_resid) > threshold_resid)[0]

    # Leverage (hat matrix diagonal)
    try:
        H = X_design @ np.linalg.pinv(X_design.T @ X_design) @ X_design.T
        leverage = np.diag(H)
        threshold_lev = 2.0 * (p + 1) / n
        high_leverage = np.where(leverage > threshold_lev)[0]
    except Exception:
        leverage = np.zeros(n)
        high_leverage = np.array([])

    return {
        "large_residuals": large_resid,
        "n_large_residuals": len(large_resid),
        "high_leverage": high_leverage,
        "n_high_leverage": len(high_leverage),
        "std_residuals": std_resid,
        "leverage": leverage,
    }


def check_residual_normality(residuals: np.ndarray) -> Dict:
    """Shapiro-Wilk test for normality of residuals."""
    n = len(residuals)
    if n < 3:
        return {"test": "Shapiro-Wilk", "statistic": np.nan, "p_value": np.nan,
                "normal": True, "n": n}
    # Shapiro-Wilk max 5000 samples
    sample = residuals[:5000] if n > 5000 else residuals
    stat, p_val = shapiro(sample)
    return {
        "test": "Shapiro-Wilk",
        "statistic": stat,
        "p_value": p_val,
        "normal": p_val > 0.05,
        "n": n,
    }


def generate_report_card(
    n: int,
    p: int,
    unusual: Dict,
    normality: Dict,
    response_name: str,
) -> List[Dict]:
    """Generate Minitab-style report card checks."""
    checks = []

    # 1. Amount of Data
    if n >= 50:
        status = "ok"
        desc = (
            f"The sample size (n = {n}) is large enough to provide a precise "
            f"estimate of the strength of the relationship."
        )
    else:
        status = "info"
        desc = (
            f"The sample size (n = {n}) is not large enough to provide a very "
            f"precise estimate of the strength of the relationship. Measures of "
            f"the strength of the relationship, such as R-Squared and R-Squared "
            f"(adjusted), can vary a great deal. To obtain a precise estimate, "
            f"larger samples (typically 50 or more) should be used for a model "
            f"of this size."
        )
    checks.append({"check": "Amount of Data", "status": status, "description": desc})

    # 2. Unusual Data
    n_lr = unusual["n_large_residuals"]
    n_hl = unusual["n_high_leverage"]
    if n_lr == 0 and n_hl == 0:
        status = "ok"
        desc = "No unusual data points were detected."
    else:
        status = "warning"
        parts = []
        if n_lr > 0:
            parts.append(
                f"Large residuals: {n_lr} data points have large residuals and "
                f"are not well fit by the equation. These points are marked in red "
                f"on the Diagnostic Report."
            )
        if n_hl > 0:
            parts.append(
                f"Unusual X values: {n_hl} data points have unusual X values, "
                f"which can strongly influence the model equation. These points "
                f"are marked in blue on the Diagnostic Report."
            )
        parts.append(
            "Because unusual data can have a strong influence on the results, "
            "try to identify the cause for their unusual nature. Correct any "
            "data entry or measurement errors. Consider removing data that are "
            "associated with special causes and redoing the analysis."
        )
        desc = " • ".join(parts)
    checks.append({"check": "Unusual Data", "status": status, "description": desc})

    # 3. Normality
    if normality["n"] >= 15:
        status = "ok"
        desc = (
            f"Because you have at least 15 data points, normality is not an "
            f"issue. If the number of data points is small and the residuals are "
            f"not normally distributed, the p-values used to determine whether "
            f"there is a significant relationship between the Xs and Y may not "
            f"be accurate."
        )
    elif normality["normal"]:
        status = "ok"
        desc = "The residuals appear to be normally distributed."
    else:
        status = "warning"
        desc = (
            "The residuals do not appear to be normally distributed. P-values "
            "may not be accurate. Consider transforming the response variable."
        )
    checks.append({"check": "Normality", "status": status, "description": desc})

    # 4. Evaluate Solutions
    checks.append({
        "check": "Evaluate Solutions",
        "status": "info",
        "description": (
            f"The Prediction and Optimization Report displays the optimal "
            f"solution, which is calculated using a numerical algorithm. It also "
            f"displays a list of alternative solutions. When evaluating these "
            f"solutions, consider the following:\n"
            f"• The feasibility of the optimal solution from a practical "
            f"perspective.\n"
            f"• The sensitivity of Y to departures from the optimal X values.\n"
            f"• The true optimal solution may not be located in the current "
            f"sample space.\n"
            f"• The alternative solutions may be more practical than the optimal "
            f"solution.\n"
            f"Once you have selected a solution, perform 20-30 confirmation "
            f"runs to validate the optimal X values."
        ),
    })

    return checks


# =============================================================================
# 6. FORMATTING HELPERS
# =============================================================================

def format_equation(
    response_name: str,
    feature_names: List[str],
    coefficients: np.ndarray,
) -> str:
    """Format model equation in Minitab style."""
    intercept = coefficients[0]
    parts = [f"{response_name} = {intercept:.4g}"]
    for name, coef in zip(feature_names[1:], coefficients[1:]):
        if abs(coef) < 1e-12:
            continue
        sign = "+" if coef >= 0 else "-"
        parts.append(f"{sign} {abs(coef):.4g} {name}")
    return " ".join(parts)


# =============================================================================
# 7. PLOTTING FUNCTIONS
# =============================================================================

def plot_model_building_sequence(steps: List[Dict]) -> go.Figure:
    """Minitab-style horizontal bar chart for forward stepwise R²(adj)."""
    if not steps:
        fig = go.Figure()
        fig.add_annotation(text="No model building steps", x=0.5, y=0.5,
                           showarrow=False, xref="paper", yref="paper")
        return fig

    labels = [f"{s['step']}" for s in steps]
    r2_vals = [s["r_squared_adj"] for s in steps]
    is_higher = [s.get("is_higher", False) for s in steps]

    fig = go.Figure()

    for i, (label, r2, higher) in enumerate(zip(labels, r2_vals, is_higher)):
        fig.add_trace(go.Bar(
            x=[r2], y=[i],
            orientation="h",
            marker_color="#4a8c3f" if not higher else "#6aad5a",
            marker_pattern_shape="/" if higher else "",
            marker_line=dict(color="#3a6e30", width=1),
            showlegend=False,
            hovertemplate=f"Step {label}: R²(adj) = {r2:.1f}%<extra></extra>",
        ))

    fig.update_layout(
        height=max(180, len(steps) * 40 + 60),
        margin=dict(l=10, r=10, t=10, b=30),
        template="plotly_white",
        xaxis=dict(title="R-Squared(adjusted) %", range=[0, 105],
                   dtick=25, showgrid=True),
        yaxis=dict(
            tickvals=list(range(len(steps))),
            ticktext=labels,
            autorange="reversed",
            showgrid=False,
        ),
        bargap=0.25,
    )
    return fig


def plot_incremental_impact(impacts: Dict[str, float]) -> go.Figure:
    """Horizontal bar chart – increase in R² per original variable."""
    names = list(impacts.keys())
    vals = list(impacts.values())

    fig = go.Figure(go.Bar(
        x=vals, y=names,
        orientation="h",
        marker_color="steelblue",
        marker_line=dict(color="#2c5f8a", width=1),
    ))
    fig.update_layout(
        height=max(140, len(names) * 45 + 50),
        margin=dict(l=10, r=10, t=10, b=30),
        template="plotly_white",
        xaxis=dict(title="Increase in R-Squared %", range=[0, max(max(vals) * 1.15, 10)],
                   showgrid=True),
        yaxis=dict(autorange="reversed", showgrid=False),
        bargap=0.35,
    )
    return fig


def plot_x_regressed_on_others(vif_r2: Dict[str, float]) -> go.Figure:
    """Horizontal bar chart – R² of each Xi regressed on all other Xj."""
    names = list(vif_r2.keys())
    vals = list(vif_r2.values())

    fig = go.Figure(go.Bar(
        x=vals, y=names,
        orientation="h",
        marker_color="steelblue",
        marker_line=dict(color="#2c5f8a", width=1),
    ))
    fig.update_layout(
        height=max(140, len(names) * 45 + 50),
        margin=dict(l=10, r=10, t=10, b=30),
        template="plotly_white",
        xaxis=dict(title="R-Squared %", range=[0, 105], showgrid=True),
        yaxis=dict(autorange="reversed", showgrid=False),
        bargap=0.35,
    )
    return fig


def plot_pvalue_bar(p_value: float, alpha: float = 0.10) -> go.Figure:
    """Minitab-style horizontal p-value bar (Yes/No)."""
    max_x = max(0.5, p_value * 1.3, alpha * 5)
    bar_len = min(p_value, max_x)
    bar_colour = "#d4a017" if p_value < alpha else "#b8c6d6"

    fig = go.Figure()
    fig.add_shape(type="rect", x0=0, x1=alpha, y0=-0.5, y1=0.5,
                  fillcolor="rgba(70,130,180,0.12)", line=dict(width=0))

    fig.add_trace(go.Bar(
        x=[bar_len], y=[0], orientation="h",
        marker_color=bar_colour,
        marker_line=dict(color="#8a7000", width=1.5),
        width=0.55, showlegend=False,
    ))
    fig.add_shape(type="line", x0=alpha, x1=alpha, y0=-0.55, y1=0.55,
                  line=dict(color="red", width=2))

    p_text = f"P < 0.001" if p_value < 0.001 else f"P = {p_value:.3f}"
    fig.add_annotation(x=bar_len, y=0, text=f"  {p_text}", showarrow=False,
                       xanchor="left", font=dict(size=12, color="black"))

    tick_vals = sorted(set([0] +
        [round(alpha * k, 3) for k in [1, 2, 4, 6, 10]
         if round(alpha * k, 3) <= max_x]))

    fig.update_layout(
        height=100, margin=dict(l=5, r=5, t=5, b=5),
        template="plotly_white",
        xaxis=dict(range=[0, max_x], tickvals=tick_vals, tickfont=dict(size=10),
                   showgrid=False),
        yaxis=dict(showticklabels=False, showgrid=False, zeroline=False,
                   range=[-0.7, 0.7]),
        annotations=[
            dict(x=0, y=-0.65, text="<b>Yes</b>", showarrow=False,
                 xref="x", yref="y", font=dict(size=11, color="#2c5f8a"),
                 xanchor="left"),
            dict(x=max_x, y=-0.65, text="<b>No</b>", showarrow=False,
                 xref="x", yref="y", font=dict(size=11, color="#8a2c2c"),
                 xanchor="right"),
        ],
    )
    return fig


def plot_scatter_panels(
    data: pd.DataFrame,
    response: str,
    predictors: List[str],
    in_model: Optional[List[str]] = None,
) -> go.Figure:
    """Y vs each X predictor in separate panels. Gray background if not in model."""
    n_pred = len(predictors)
    if n_pred == 0:
        fig = go.Figure()
        fig.add_annotation(text="No predictors", x=0.5, y=0.5,
                           showarrow=False, xref="paper", yref="paper")
        return fig

    cols = min(n_pred, 4)
    rows = (n_pred + cols - 1) // cols

    fig = make_subplots(rows=rows, cols=cols,
                        subplot_titles=predictors,
                        horizontal_spacing=0.08,
                        vertical_spacing=0.12)

    y_vals_all = pd.to_numeric(data[response], errors="coerce").dropna()
    y_pad = (y_vals_all.max() - y_vals_all.min()) * 0.08

    for i, pred in enumerate(predictors):
        r = i // cols + 1
        c = i % cols + 1
        y_vals = pd.to_numeric(data[response], errors="coerce")
        x_vals = pd.to_numeric(data[pred], errors="coerce")

        is_in = in_model is None or pred in in_model

        fig.add_trace(go.Scatter(
            x=x_vals, y=y_vals,
            mode="markers",
            marker=dict(color="#2c5f8a", size=5),
            showlegend=False,
        ), row=r, col=c)

        if not is_in:
            x_pad = (x_vals.max() - x_vals.min()) * 0.15
            # Use paper-relative shape via axis range
            ax_num = i + 1
            x_key = f"xaxis{ax_num}" if ax_num > 1 else "xaxis"
            y_key = f"yaxis{ax_num}" if ax_num > 1 else "yaxis"
            fig.update_layout(**{
                x_key: dict(range=[x_vals.min() - x_pad, x_vals.max() + x_pad]),
                y_key: dict(range=[y_vals.min() - y_pad, y_vals.max() + y_pad]),
            })
            fig.add_shape(
                type="rect",
                x0=x_vals.min() - x_pad * 2,
                x1=x_vals.max() + x_pad * 2,
                y0=y_vals.min() - y_pad * 2,
                y1=y_vals.max() + y_pad * 2,
                fillcolor="rgba(200,200,200,0.35)",
                layer="below", line=dict(width=0),
                row=r, col=c,
            )

    fig.update_layout(
        height=250 * rows,
        template="plotly_white",
        margin=dict(l=50, r=20, t=40, b=30),
        showlegend=False,
    )
    fig.update_yaxes(title_text=response, col=1)
    return fig


def plot_diagnostic_report(
    result: Dict,
    unusual: Dict,
    response_name: str = "Y",
) -> go.Figure:
    """
    Minitab-style diagnostic report:
      Top-left:  Residuals vs Fitted (with unusual points coloured)
      Top-right: Normal probability plot of residuals
      Bot-left:  Histogram of residuals
      Bot-right: Residuals vs observation order

    Large residuals in RED, high-leverage points in BLUE, both in PURPLE.
    """
    y_pred = result["y_pred"]
    residuals = result["residuals"]
    n = len(residuals)
    obs_order = np.arange(1, n + 1)

    std_res = unusual.get("std_residuals", residuals / np.std(residuals, ddof=1))
    large_idx = set(unusual.get("large_residuals", []))
    lever_idx = set(unusual.get("high_leverage", []))

    # Colour per point
    colours = []
    for i in range(n):
        if i in large_idx and i in lever_idx:
            colours.append("#9b59b6")  # purple — both
        elif i in large_idx:
            colours.append("#e74c3c")  # red — large residual
        elif i in lever_idx:
            colours.append("#3498db")  # blue — high leverage
        else:
            colours.append("#333333")  # normal

    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=[
            "Residuals vs Fitted Values",
            "Normal Probability Plot",
            "Histogram of Residuals",
            "Residuals vs Order",
        ],
        horizontal_spacing=0.1,
        vertical_spacing=0.14,
    )

    # --- (1,1) Residuals vs Fitted ---
    fig.add_trace(go.Scatter(
        x=y_pred, y=residuals,
        mode="markers",
        marker=dict(color=colours, size=5),
        showlegend=False,
        hovertemplate="Fitted: %{x:.2f}<br>Residual: %{y:.2f}<extra></extra>",
    ), row=1, col=1)
    fig.add_hline(y=0, line=dict(color="red", width=1, dash="dash"), row=1, col=1)
    fig.update_xaxes(title_text="Fitted Value", row=1, col=1)
    fig.update_yaxes(title_text="Residual", row=1, col=1)

    # --- (1,2) Normal probability plot ---
    sorted_res = np.sort(residuals)
    n_pts = len(sorted_res)
    theoretical_q = stats.norm.ppf((np.arange(1, n_pts + 1) - 0.375) / (n_pts + 0.25))
    fig.add_trace(go.Scatter(
        x=theoretical_q, y=sorted_res,
        mode="markers",
        marker=dict(color="#2c5f8a", size=4),
        showlegend=False,
    ), row=1, col=2)
    # Reference line
    slope = np.std(residuals, ddof=1)
    intercept = np.mean(residuals)
    q_range = np.array([theoretical_q.min(), theoretical_q.max()])
    fig.add_trace(go.Scatter(
        x=q_range, y=intercept + slope * q_range,
        mode="lines",
        line=dict(color="red", width=1.5),
        showlegend=False,
    ), row=1, col=2)
    fig.update_xaxes(title_text="Theoretical Quantile", row=1, col=2)
    fig.update_yaxes(title_text="Residual", row=1, col=2)

    # --- (2,1) Histogram ---
    fig.add_trace(go.Histogram(
        x=residuals,
        marker_color="steelblue",
        marker_line=dict(color="white", width=0.5),
        showlegend=False,
    ), row=2, col=1)
    fig.update_xaxes(title_text="Residual", row=2, col=1)
    fig.update_yaxes(title_text="Frequency", row=2, col=1)

    # --- (2,2) Residuals vs Order ---
    fig.add_trace(go.Scatter(
        x=obs_order, y=residuals,
        mode="lines+markers",
        marker=dict(color=colours, size=4),
        line=dict(color="#aaa", width=0.8),
        showlegend=False,
        hovertemplate="Obs: %{x}<br>Residual: %{y:.2f}<extra></extra>",
    ), row=2, col=2)
    fig.add_hline(y=0, line=dict(color="red", width=1, dash="dash"), row=2, col=2)
    fig.update_xaxes(title_text="Observation Order", row=2, col=2)
    fig.update_yaxes(title_text="Residual", row=2, col=2)

    fig.update_layout(
        height=600,
        template="plotly_white",
        margin=dict(l=60, r=30, t=40, b=40),
        showlegend=False,
    )
    return fig


def export_results_to_excel(
    result: Dict,
    steps: List[Dict],
    impacts: Dict[str, float],
    vif_r2: Dict[str, float],
    unusual: Dict,
    normality: Dict,
    report_card: List[Dict],
    equation: str,
    response_name: str,
    filepath: str,
):
    """Export all regression results to a multi-sheet Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    title_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def _add_title(ws, title, subtitle=""):
        ws.merge_cells("A1:F1")
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        if subtitle:
            ws.merge_cells("A2:F2")
            ws["A2"] = subtitle
            ws["A2"].font = Font(bold=True, color="FFFFFF", size=11)
            ws["A2"].fill = title_fill
            ws["A2"].alignment = Alignment(horizontal="center")

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    _add_title(ws, f"Multiple Regression for {response_name}", "Summary Report")

    ws["A4"] = "Model Equation:"
    ws["A4"].font = bold
    ws.merge_cells("B4:F4")
    ws["B4"] = equation

    ws["A6"] = "R²"
    ws["B6"] = result["r_squared"]
    ws["A7"] = "R²(adj)"
    ws["B7"] = result["r_squared_adj"]
    ws["A8"] = "S"
    ws["B8"] = result["s"]
    ws["A9"] = "F-statistic"
    ws["B9"] = result["ols_stats"]["f_statistic"]
    ws["A10"] = "F p-value"
    ws["B10"] = result["ols_stats"]["f_p_value"]
    ws["A11"] = "N"
    ws["B11"] = result["n"]
    ws["A12"] = "Predictors"
    ws["B12"] = result["p"]
    for r_idx in range(6, 13):
        ws[f"A{r_idx}"].font = bold

    # ---- Sheet 2: Coefficients ----
    ws2 = wb.create_sheet("Coefficients")
    _add_title(ws2, f"Coefficient Table", response_name)

    headers = ["Term", "Coefficient", "SE", "t-Value", "p-Value"]
    for j, h in enumerate(headers):
        cell = ws2.cell(row=4, column=j + 1, value=h)
        cell.font = bold
        cell.fill = header_fill

    names = result["feature_names"]
    coefs = result["coefficients"]
    se = result["ols_stats"]["se"]
    t_vals = result["ols_stats"]["t_values"]
    p_vals = result["ols_stats"]["p_values"]

    for i in range(len(names)):
        ws2.cell(row=5 + i, column=1, value=names[i])
        ws2.cell(row=5 + i, column=2, value=round(float(coefs[i]), 6))
        ws2.cell(row=5 + i, column=3, value=round(float(se[i]), 6))
        ws2.cell(row=5 + i, column=4, value=round(float(t_vals[i]), 4))
        ws2.cell(row=5 + i, column=5, value=round(float(p_vals[i]), 6))

    # ---- Sheet 3: Model Building ----
    ws3 = wb.create_sheet("Model Building")
    _add_title(ws3, "Model Building Sequence", response_name)

    headers3 = ["Step", "Change", "Step P", "Final P", "R²(adj) %"]
    for j, h in enumerate(headers3):
        cell = ws3.cell(row=4, column=j + 1, value=h)
        cell.font = bold
        cell.fill = header_fill

    for i, s in enumerate(steps):
        ws3.cell(row=5 + i, column=1, value=s["step"])
        ws3.cell(row=5 + i, column=2, value=s["change"])
        ws3.cell(row=5 + i, column=3, value=round(s["step_p"], 4) if s["step_p"] is not None else "")
        ws3.cell(row=5 + i, column=4, value=round(s["final_p"], 4) if s["final_p"] is not None else "")
        ws3.cell(row=5 + i, column=5, value=round(s["r_squared_adj"], 2))

    # Incremental impact
    row_start = 5 + len(steps) + 2
    ws3.cell(row=row_start, column=1, value="Incremental Impact of X Variables").font = bold
    ws3.cell(row=row_start + 1, column=1, value="Variable").font = bold
    ws3.cell(row=row_start + 1, column=2, value="Increase in R² %").font = bold
    ws3.cell(row=row_start + 1, column=1).fill = header_fill
    ws3.cell(row=row_start + 1, column=2).fill = header_fill
    for i, (k, v) in enumerate(impacts.items()):
        ws3.cell(row=row_start + 2 + i, column=1, value=k)
        ws3.cell(row=row_start + 2 + i, column=2, value=round(v, 4))

    # VIF
    row_start2 = row_start + 2 + len(impacts) + 2
    ws3.cell(row=row_start2, column=1, value="Each X Regressed on All Others").font = bold
    ws3.cell(row=row_start2 + 1, column=1, value="Variable").font = bold
    ws3.cell(row=row_start2 + 1, column=2, value="R² %").font = bold
    ws3.cell(row=row_start2 + 1, column=1).fill = header_fill
    ws3.cell(row=row_start2 + 1, column=2).fill = header_fill
    for i, (k, v) in enumerate(vif_r2.items()):
        ws3.cell(row=row_start2 + 2 + i, column=1, value=k)
        ws3.cell(row=row_start2 + 2 + i, column=2, value=round(v, 4))

    # ---- Sheet 4: Diagnostics ----
    ws4 = wb.create_sheet("Diagnostics")
    _add_title(ws4, "Diagnostic Report", response_name)

    red_font = Font(color="CC0000", bold=True)
    blue_font = Font(color="0000CC", bold=True)

    diag_headers = ["Obs", "Fitted", "Residual", "Std Residual", "Leverage", "Flag"]
    for j, h in enumerate(diag_headers):
        cell = ws4.cell(row=4, column=j + 1, value=h)
        cell.font = bold
        cell.fill = header_fill

    y_pred = result["y_pred"]
    residuals = result["residuals"]
    std_res = unusual.get("std_residuals", np.zeros(len(residuals)))
    leverage = unusual.get("leverage", np.zeros(len(residuals)))
    large_set = set(unusual.get("large_residuals", []))
    lever_set = set(unusual.get("high_leverage", []))

    for i in range(len(residuals)):
        row_num = 5 + i
        ws4.cell(row=row_num, column=1, value=i + 1)
        ws4.cell(row=row_num, column=2, value=round(float(y_pred[i]), 4))
        ws4.cell(row=row_num, column=3, value=round(float(residuals[i]), 4))
        ws4.cell(row=row_num, column=4, value=round(float(std_res[i]), 4))
        ws4.cell(row=row_num, column=5, value=round(float(leverage[i]), 4))

        flags = []
        if i in large_set:
            flags.append("Large Residual")
        if i in lever_set:
            flags.append("High Leverage")
        flag_text = "; ".join(flags) if flags else ""
        cell = ws4.cell(row=row_num, column=6, value=flag_text)

        # Colour the row
        if i in large_set and i in lever_set:
            for col in range(1, 7):
                ws4.cell(row=row_num, column=col).font = Font(color="9B59B6", bold=True)
        elif i in large_set:
            for col in range(1, 7):
                ws4.cell(row=row_num, column=col).font = red_font
        elif i in lever_set:
            for col in range(1, 7):
                ws4.cell(row=row_num, column=col).font = blue_font

    # ---- Sheet 5: Report Card ----
    ws5 = wb.create_sheet("Report Card")
    _add_title(ws5, "Report Card", response_name)

    for j, h in enumerate(["Check", "Status", "Description"]):
        cell = ws5.cell(row=4, column=j + 1, value=h)
        cell.font = bold
        cell.fill = header_fill

    for i, c in enumerate(report_card):
        ws5.cell(row=5 + i, column=1, value=c["check"]).font = bold
        ws5.cell(row=5 + i, column=2, value=c["status"].upper())
        ws5.cell(row=5 + i, column=3, value=c["description"])

    # Auto-width columns
    for ws_item in [ws, ws2, ws3, ws4, ws5]:
        for col in ws_item.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_item.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(filepath)
    return filepath
